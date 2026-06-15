import os
import requests
import pandas as pd
from dotenv import load_dotenv
import anthropic
import json
import argparse
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Alignment

try:
    from ddgs import DDGS
except ImportError:
    DDGS = None
    print("⚠ ddgs not installed. Run: pip install ddgs")

load_dotenv()

with open("master_resume.json", "r") as f:
    MASTER_RESUME = json.load(f)

NEWS_API_KEY = os.getenv("NEWS_API_KEY")
client = anthropic.Anthropic()

# ── Data fetchers ──────────────────────────────────────────────────────────

def get_company_web_summary(company_name):
    """Pull a company overview by combining the top 3 DuckDuckGo web results."""
    if DDGS is None:
        return ""
    try:
        results = list(DDGS().text(company_name, max_results=3))
        if not results:
            return ""
        lines = []
        for r in results:
            title = (r.get("title") or "").strip()
            url = (r.get("href") or "").strip()
            snippet = (r.get("body") or "").strip()
            lines.append(f"{title}\n{url}\n{snippet}")
        return "\n\n".join(lines)
    except Exception:
        return ""

def get_news_articles(company_name, num_articles=3):
    """Pull recent news from NewsAPI: display line + article URL per item."""
    try:
        url = "https://newsapi.org/v2/everything"
        params = {
            "q": f'"{company_name}"',
            "sortBy": "publishedAt",
            "pageSize": num_articles,
            "language": "en",
            "apiKey": NEWS_API_KEY
        }
        response = requests.get(url, params=params, timeout=10)
        print(f"  News API status: {response.status_code}", flush=True)

        if response.status_code == 200:
            articles = response.json().get("articles", [])
            print(f"  Articles returned: {len(articles)}", flush=True)
            out = []
            for a in articles:
                title = a.get("title", "") or ""
                published = (a.get("publishedAt") or "")[:10]
                source = (a.get("source") or {}).get("name", "") or ""
                line = f"{published} [{source}]: {title}"
                article_url = (a.get("url") or "").strip()
                out.append({"line": line, "url": article_url})
            return out
        else:
            print(f"  News API error: {response.json().get('message', 'Unknown error')}", flush=True)
        return []
    except Exception as e:
        print(f"  News fetch error: {e}", flush=True)
        return []


def get_company_intel(company_name, job_title, news_articles=None):
    """Use Claude to synthesize all available data into a structured brief."""
    web_summary = get_company_web_summary(company_name)
    if news_articles is None:
        news_articles = get_news_articles(company_name)
    news_lines = [x["line"] for x in news_articles] if news_articles else []
    news_text = "\n".join(news_lines) if news_lines else "No recent news found."

    print(f"  Web search data: {'Found' if web_summary else 'Not found'}", flush=True)
    print(f"  News headlines: {len(news_articles)} found", flush=True)

    prompt = f"""
You are a business intelligence analyst. Based on the information below,
provide a structured company brief for a job seeker considering applying
to this company.

COMPANY: {company_name}
JOB TITLE THEY ARE CONSIDERING: {job_title}

WEB SEARCH RESULTS:
{web_summary if web_summary else "No web search data found — use general knowledge."}

RECENT NEWS HEADLINES:
{news_text}

Return ONLY a JSON object in this exact format — no other text:
{{
  "industry": "primary industry",
  "hq_location": "city, state/country",
  "company_size": "estimated employee count or range",
  "founded": "year founded if known",
  "market_cap": "market cap or revenue estimate if known, or Private",
  "description": "2-3 sentence plain english summary of what the company does",
  "stability": "Strong / Stable / Uncertain / Unknown",
  "growth_trend": "Growing / Stable / Declining / Unknown",
  "news_1": "most recent relevant headline with date, or No recent news found",
  "news_2": "second most recent headline with date, or empty string",
  "news_3": "third most recent headline with date, or empty string",
  "recommendation": "1-2 sentence assessment of this company as an employer given the job seeker's target role"
}}
"""

    message = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=1024,
        messages=[{"role": "user", "content": prompt}]
    )

    response_text = message.content[0].text
    start = response_text.find("{")
    end = response_text.rfind("}") + 1
    return json.loads(response_text[start:end])


def get_resume_match(job_description, master_resume):
    """Use Claude to score how well the master resume matches a job description.

    Returns a dict with match_score (int 0-100) and match_notes (str).
    """
    prompt = f"""You are evaluating how well a candidate's resume matches a job description.

MASTER RESUME (JSON — bullets, tags, and strength scores):
{json.dumps(master_resume, indent=2)}

JOB DESCRIPTION:
{job_description}

Analyze the match. Return ONLY a JSON object in this exact format — no other text:
{{
  "match_score": <integer 0-100>,
  "match_notes": "<2-3 specific strengths that match this role, then 1-2 gaps. Plain text, no bullet symbols, pipe-separated. Example: Strong program delivery experience matches PM requirements | PMP cert directly relevant | No cloud infrastructure background mentioned in resume>"
}}

Scoring guide:
- 80-100: Strong match. Most key requirements covered, relevant industry/domain.
- 60-79: Good match. Core skills align, some gaps.
- 40-59: Partial match. Transferable skills but meaningful gaps.
- Below 40: Weak match. Major requirements unmet.
"""

    message = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=1024,
        messages=[{"role": "user", "content": prompt}]
    )

    response_text = message.content[0].text
    start = response_text.find("{")
    end = response_text.rfind("}") + 1
    return json.loads(response_text[start:end])


def _normalize_founded_value(raw) -> str:
    """Avoid Excel float year artifacts when saving (e.g. store 1995 not 1995.0)."""
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return ""
    s = str(raw).strip()
    if not s or s.lower() == "nan":
        return ""
    try:
        f = float(s.replace(",", ""))
        if f == int(f) and 1000 <= abs(f) <= 9999:
            return str(int(f))
    except ValueError:
        pass
    return s


def _normalize_match_score(raw):
    """Keep the match score as an int when possible; blank otherwise (no '50.0')."""
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return ""
    s = str(raw).strip()
    if not s or s.lower() == "nan":
        return ""
    try:
        return int(float(s))
    except ValueError:
        return ""


ENRICHMENT_TEXT_COLUMNS = [
    "Industry", "HQ Location", "Company Size", "Founded",
    "Market Cap", "Description", "Stability", "Growth Trend",
    "News 1", "News 2", "News 3",
    "News 1 URL", "News 2 URL", "News 3 URL",
    "Recommendation", "Enriched Date",
    "Match Score", "Match Notes",
]


def _prepare_enrichment_columns(df: pd.DataFrame) -> None:
    """Excel often loads Founded as float64 (1995.0). Coerce text columns to object
    so re-enrichment can assign string years like '1995' without dtype errors."""
    for col in ENRICHMENT_TEXT_COLUMNS:
        if col not in df.columns:
            continue
        if col == "Founded":
            df[col] = df[col].apply(_normalize_founded_value)
        elif col == "Match Score":
            # Coerce to int where possible (not string) so 50 stays 50, not "50.0".
            df[col] = df[col].apply(_normalize_match_score)
        else:
            df[col] = df[col].fillna("").astype(str)
        df[col] = df[col].astype(object)

# ── Text wrap columns in Excel ─────────────────────────────────────────────

def apply_formatting(excel_file):
    """Apply text wrapping to description and recommendation columns."""
    wb = load_workbook(excel_file)
    ws = wb.active

    # Find columns to wrap
    wrap_columns = ["Description", "Recommendation", "News 1", "News 2", "News 3"]
    col_indices = {}

    for cell in ws[1]:  # Header row
        if cell.value in wrap_columns:
            col_indices[cell.column] = cell.value

    # Apply wrap to those columns
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column in col_indices:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Set column widths
    col_widths = {
        "Description":    50,
        "Recommendation": 50,
        "News 1":         60,
        "News 2":         60,
        "News 3":         60,
    }

    for cell in ws[1]:
        if cell.value in col_widths:
            ws.column_dimensions[cell.column_letter].width = col_widths[cell.value]

    # Set row height for data rows
    for row_num in range(2, ws.max_row + 1):
        ws.row_dimensions[row_num].height = 80

    wb.save(excel_file)
    print("Formatting applied.")

# ── Main enrichment loop ───────────────────────────────────────────────────

def enrich_jobs(excel_file=os.path.join("output", "job_listings.xlsx")):
    print(f"Loading {excel_file}...", flush=True)
    df = pd.read_excel(excel_file)

    # Find rows marked as Interested
    interested = df[df["Status"].str.strip().str.lower() == "interested"]

    if interested.empty:
        print("No jobs marked as 'Interested' found.", flush=True)
        print(f"Open {excel_file}, change Status to 'Interested' for jobs you want, then run again.", flush=True)
        return

    print(f"Found {len(interested)} jobs marked as Interested. Enriching...", flush=True)

    # Add new columns if they don't exist
    new_cols = ["Industry", "HQ Location", "Company Size", "Founded",
                "Market Cap", "Description", "Stability", "Growth Trend",
                "News 1", "News 2", "News 3",
                "News 1 URL", "News 2 URL", "News 3 URL",
                "Recommendation", "Enriched Date",
                "Match Score", "Match Notes"]
    for col in new_cols:
        if col not in df.columns:
            df[col] = ""

    _prepare_enrichment_columns(df)

    # Enrich each interested row
    for idx, row in interested.iterrows():
        company = row["Company"]
        title = row["Title"]
        print(f"\nEnriching {company} via DuckDuckGo / NewsAPI...", flush=True)

        try:
            news_articles = get_news_articles(company)
            intel = get_company_intel(company, title, news_articles=news_articles)

            df.at[idx, "Industry"]       = intel.get("industry", "")
            df.at[idx, "HQ Location"]    = intel.get("hq_location", "")
            df.at[idx, "Company Size"]   = intel.get("company_size", "")
            df.at[idx, "Founded"]        = _normalize_founded_value(intel.get("founded", ""))
            df.at[idx, "Market Cap"]     = intel.get("market_cap", "")
            df.at[idx, "Description"]    = intel.get("description", "")
            df.at[idx, "Stability"]      = intel.get("stability", "")
            df.at[idx, "Growth Trend"]   = intel.get("growth_trend", "")
            df.at[idx, "News 1"]         = intel.get("news_1", "")
            df.at[idx, "News 2"]         = intel.get("news_2", "")
            df.at[idx, "News 3"]         = intel.get("news_3", "")
            for i in range(3):
                col_url = f"News {i + 1} URL"
                u = news_articles[i]["url"] if i < len(news_articles) else ""
                df.at[idx, col_url] = u or ""
            df.at[idx, "Recommendation"] = intel.get("recommendation", "")
            df.at[idx, "Enriched Date"]  = str(date.today())

            print(f"  ✓ Done — {intel.get('stability')} / {intel.get('growth_trend')}", flush=True)

            # ── Resume match score ──────────────────────────────────────
            # Only score when there is a real job description and the row
            # has not already been scored (avoid re-computing on re-runs).
            # A failure here must not abort the enrichment that just succeeded.
            jd_raw = row.get("Job Description", "")
            job_description = (
                "" if (jd_raw is None or (isinstance(jd_raw, float) and pd.isna(jd_raw)))
                else str(jd_raw)
            )
            existing_score = df.at[idx, "Match Score"]
            already_scored = not (
                existing_score is None
                or existing_score == ""
                or (isinstance(existing_score, float) and pd.isna(existing_score))
            )

            if already_scored:
                print(f"  Match score already present ({existing_score}) — skipping.", flush=True)
            elif len(job_description) > 200:
                try:
                    intel_match = get_resume_match(job_description, MASTER_RESUME)
                    df.at[idx, "Match Score"] = intel_match.get("match_score", "")
                    df.at[idx, "Match Notes"] = intel_match.get("match_notes", "")
                    print(f"  ✓ Resume match: {intel_match.get('match_score')}/100", flush=True)
                except Exception as e:
                    print(f"  ✗ Match score error: {e}", flush=True)
            else:
                print("  ⚠ No job description — match score skipped", flush=True)

        except Exception as e:
            print(f"  ✗ Error enriching {company}: {e}", flush=True)

    # Save to Excel
    df.to_excel(excel_file, index=False)
    print(f"\nData saved to {excel_file}", flush=True)

    # Apply formatting
    apply_formatting(excel_file)
    print(f"Enrichment complete. Open {excel_file} to review.", flush=True)

# ── Run ───────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Enrich Interested jobs in a listings Excel file")
    parser.add_argument(
        "excel_file",
        nargs="?",
        default=os.path.join("output", "job_listings.xlsx"),
        help="Excel file to enrich (default: output/job_listings.xlsx)",
    )
    args = parser.parse_args()
    enrich_jobs(args.excel_file)
